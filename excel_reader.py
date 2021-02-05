import os
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from prettytable import PrettyTable

class SalaryFileReader:
    def __init__(self, file_path, sheet_name, config) :
        if not os.path.exists(file_path):
            raise IOError('File not found')
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.titles = None
        self.values = None
        self.user_value_map = None
        self.table_view = None
        self.config = config
        self.has_bye_people = False
        pass

    def set_sheet_name(self, sheet_name):
        self.sheet_name = sheet_name

    def load(self):
        self.table_view = None
        self.values = None
        self.user_value_map = None
        self.titles = None

        if not os.path.exists(self.file_path):
            raise IOError('File not found')
        wb = load_workbook(self.file_path)
        # print(wb.sheetnames)
        if self.sheet_name not in wb:
            print('Sheet [{0}] not found in [{1}]'.format(self.sheet_name, ','.join(wb.sheetnames)))
            raise SheetNotFound(wb.sheetnames)

        sheet = wb[self.sheet_name]
        # print(sheet.dimensions)
        # print('%d:%d' % (sheet.max_column, sheet.max_row))
        start_from_row = 1
        column_A = [value[0].value for value in sheet['A1': 'A10']]
        for start_from_row in range(1, len(column_A)):
            if column_A[start_from_row - 1] == '序号':
                break
            if start_from_row == len(column_A) - 1:
                print('表格中没有找到有效数据')
                raise ValueError

        titles = sheet['A%d' % start_from_row: '%s%d' % (get_column_letter(sheet.max_column), start_from_row + 1)]
        title_strs = []

        last_cell = None
        for cell in titles[0]:
            if cell.value is not None:
                last_cell = cell.value
            if cell.value in self.config['no_concat_titles']:
                last_cell = None
            next_cell = titles[1][cell.column - 1]
            if next_cell.value is None:
                if cell.value is not None:
                    title_strs.append(cell.value)
            else:
                if last_cell:
                    title_strs.append('%s-%s' % (last_cell, next_cell.value))
                else:
                    title_strs.append(next_cell.value)
            pass

        values = sheet['A%d' % (start_from_row + 2): '%s%d' % (get_column_letter(sheet.max_column), sheet.max_row)]
        values = [[round(cell.value, 2) if type(cell.value) == float
                   else cell.value.date() if type(cell.value) == datetime
                   else cell.value
                   for cell in row] for row in values]
        values = filter(lambda r: r[0] is not None and r[0] != '合计', values)
        self.titles = list(title_strs)
        self.values = list(values)
        self.values_to_map()

    def get_info(self, seq):
        return self.user_value_map[seq] if seq in self.user_value_map else None

    def get_title_map(self):
        title_map = {}
        for key in self.config:
            title = self.config[key]
            if isinstance(title, list):
                title_map[key] = []
                for stitle in title:
                    if stitle in self.titles:
                        title_map[key].append(self.titles.index(stitle))
                    else:
                        title_map[key].append(None)
            elif title.endswith('{}'):
                title_map[key] = []
                title = title[: len(title) - 2]
                for index in range(0, len(self.titles)):
                    if self.titles[index] is None:
                        print('NONE %d' % index)
                    elif self.titles[index].startswith(title):
                        title_map[key].append(index)
                    pass
            else:
                title_map[key] = self.titles.index(title)
        return title_map

    def values_to_map(self):
        title_map = self.get_title_map()
        user_value_map = {}
        for row in self.values:
            result = {}
            for key in title_map:
                position = title_map[key]
                if isinstance(position, list):
                    values = [row[p] if p is not None else 0.0 for p in position]
                    result[key] = [value if value is not None else 0.0 for value in values]
                else:
                    result[key] = row[position]
            if result['out_day']:
                self.has_bye_people = True
            if result['salary'] is None and result['salary2']:
                result['salary'] = result['salary2']
            mi = 2
            result['money2'] = ''
            result['money2amount'] = ''
            result['money3'] = ''
            result['money3amount'] = ''
            result['money4'] = ''
            result['money4amount'] = ''
            result['baoxian_money'] = round(sum(result['baoxian_moneys']), 2)
            for index in range(0, len(result['additions'])):
                money = result['additions'][index]
                if money is None or money == 0:
                    continue
                result['money%d' % mi] = self.config['additions'][index]
                result['money%damount' % mi] = money
                mi += 1
            user_value_map[result['seq']] = result
        self.user_value_map = user_value_map

    def table(self, title_show):
        if self.table_view:
            return self.table_view
        title_map = self.get_title_map()
        title_show_indexes = list(map(lambda i: title_map[i], title_show))
        x = PrettyTable()
        x.field_names = [self.titles[i] for i in title_show_indexes]
        for row in self.values:
            x.add_row(['' if row[i] is None else row[i] for i in title_show_indexes])
        self.table_view = x.get_string()
        return self.table_view
        pass


class SheetNotFound(ValueError):
    def __init__(self, names):
        self.sheet_names = names
    pass
